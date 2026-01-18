from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, File, UploadFile
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from supabase import create_client, Client
from passlib.context import CryptContext
from jose import JWTError, jwt
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, EmailStr, Field, ConfigDict
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import resend

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')
load_dotenv(ROOT_DIR.parent.parent / '.env')

SUPABASE_URL = os.environ.get('VITE_SUPABASE_URL') or os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('VITE_SUPABASE_SUPABASE_ANON_KEY') or os.environ.get('SUPABASE_SERVICE_ROLE_KEY') or os.environ.get('SUPABASE_ANON_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("Supabase credentials not found in environment variables")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30 * 24 * 60

RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

app = FastAPI()
api_router = APIRouter(prefix="/api")


class UserBase(BaseModel):
    username: str
    email: EmailStr
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    role: str = "user"
    sub_role: Optional[str] = None
    department: Optional[str] = None
    team: Optional[str] = None
    pillar: Optional[str] = None
    manager: Optional[str] = None
    approved_pillars: Optional[List[str]] = []
    approved_departments: Optional[List[str]] = []


class UserCreate(UserBase):
    password: str


class UserPasswordChange(BaseModel):
    current_password: str
    new_password: str


class ForgotPasswordRequest(BaseModel):
    email: EmailStr


class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str


class SubRoleSelection(BaseModel):
    sub_role: str


class UserLogin(BaseModel):
    username: str
    password: str


class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    created_at: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: User


class IdeaBase(BaseModel):
    pillar: str
    title: str
    improvement_type: str
    current_process: str
    suggested_solution: str
    benefits: str
    target_completion: str
    department: Optional[str] = None
    team: Optional[str] = None


class IdeaCreate(IdeaBase):
    pass


class CIEvaluation(BaseModel):
    is_quick_win: bool
    complexity_level: Optional[str] = None
    savings_type: Optional[str] = None
    cost_savings: Optional[float] = None
    time_saved_hours: Optional[float] = None
    time_saved_minutes: Optional[float] = None
    evaluation_notes: Optional[str] = None
    assigned_to_tech: Optional[bool] = False
    tech_person_name: Optional[str] = None


class BestIdeaSelection(BaseModel):
    idea_id: str
    is_best_idea: bool


class Idea(IdeaBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    idea_number: str
    status: str
    submitted_by: str
    submitted_by_username: str
    assigned_approver: Optional[str] = None
    assigned_approver_username: Optional[str] = None
    created_at: str
    updated_at: str
    is_quick_win: Optional[bool] = None
    complexity_level: Optional[str] = None
    savings_type: Optional[str] = None
    cost_savings: Optional[float] = None
    time_saved_hours: Optional[float] = None
    time_saved_minutes: Optional[float] = None
    evaluation_notes: Optional[str] = None
    assigned_to_tech: Optional[bool] = False
    tech_person_name: Optional[str] = None
    is_best_idea: Optional[bool] = False
    evaluated_by: Optional[str] = None
    evaluated_by_username: Optional[str] = None
    evaluated_at: Optional[str] = None
    is_evaluated: Optional[bool] = False


class CommentBase(BaseModel):
    comment_text: str


class Comment(CommentBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    idea_id: str
    user_id: str
    username: str
    created_at: str


class IdeaAction(BaseModel):
    comment: Optional[str] = None


class DepartmentBase(BaseModel):
    name: str
    pillar: str


class Department(DepartmentBase):
    model_config = ConfigDict(extra="ignore")
    id: str


class PillarBase(BaseModel):
    name: str


class Pillar(PillarBase):
    model_config = ConfigDict(extra="ignore")
    id: str


class TeamBase(BaseModel):
    name: str
    pillar: str
    department: str


class Team(TeamBase):
    model_config = ConfigDict(extra="ignore")
    id: str


class TechPersonBase(BaseModel):
    name: str
    email: Optional[str] = None
    specialization: Optional[str] = None


class TechPerson(TechPersonBase):
    model_config = ConfigDict(extra="ignore")
    id: str


class DashboardStats(BaseModel):
    total_ideas: int
    pending_ideas: int
    approved_ideas: int
    declined_ideas: int
    revision_requested_ideas: int
    my_ideas: int


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def create_reset_token(email: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(hours=1)
    to_encode = {"email": email, "exp": expire, "type": "password_reset"}
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def verify_reset_token(token: str) -> str:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("email")
        token_type: str = payload.get("type")
        if email is None or token_type != "password_reset":
            raise HTTPException(status_code=400, detail="Invalid reset token")
        return email
    except JWTError:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")

        result = supabase.table("profiles").select("*").eq("id", user_id).maybeSingle().execute()
        user = result.data
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")


async def get_admin_user(current_user: dict = Depends(get_current_user)) -> dict:
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


async def send_email_async(recipient_email: str, subject: str, html_content: str):
    if not RESEND_API_KEY:
        logging.warning(f"Email not sent (no API key): {subject} to {recipient_email}")
        return False

    params = {
        "from": SENDER_EMAIL,
        "to": [recipient_email],
        "subject": subject,
        "html": html_content
    }

    try:
        result = await asyncio.to_thread(resend.Emails.send, params)
        logging.info(f"Email sent successfully: {subject} to {recipient_email}, ID: {result}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email to {recipient_email}: {str(e)}")
        return False


async def generate_idea_number() -> str:
    result = supabase.table("ideas").select("id", count="exact").execute()
    count = result.count or 0
    return f"EYE-{str(count + 1).zfill(5)}"


@api_router.get("/health")
async def health():
    return {"status": "healthy", "service": "Philtech Eye-dea API"}


@api_router.get("/public/pillars", response_model=List[Pillar])
async def get_public_pillars():
    result = supabase.table("pillars").select("*").execute()
    return [Pillar(id=str(p["id"]), name=p["name"]) for p in result.data]


@api_router.get("/public/departments", response_model=List[Department])
async def get_public_departments(pillar: Optional[str] = None):
    query = supabase.table("departments").select("*")
    if pillar:
        query = query.eq("pillar", pillar)
    result = query.execute()
    return [Department(id=str(d["id"]), name=d["name"], pillar=d["pillar"]) for d in result.data]


@api_router.get("/public/teams", response_model=List[Team])
async def get_public_teams(pillar: Optional[str] = None, department: Optional[str] = None):
    query = supabase.table("teams").select("*")
    if pillar:
        query = query.eq("pillar", pillar)
    if department:
        query = query.eq("department", department)
    result = query.execute()
    return [Team(id=str(t["id"]), name=t["name"], pillar=t["pillar"], department=t["department"]) for t in result.data]


@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    existing = supabase.table("profiles").select("id").eq("username", user_data.username).maybeSingle().execute()
    if existing.data:
        raise HTTPException(status_code=400, detail="Username already exists")

    existing_email = supabase.table("profiles").select("id").eq("email", user_data.email).maybeSingle().execute()
    if existing_email.data:
        raise HTTPException(status_code=400, detail="Email already exists")

    try:
        auth_result = supabase.auth.sign_up({
            "email": user_data.email,
            "password": user_data.password,
            "options": {
                "data": {
                    "username": user_data.username,
                    "role": user_data.role
                }
            }
        })

        if not auth_result.user:
            raise HTTPException(status_code=400, detail="Failed to create user")

        user_id = auth_result.user.id

        profile_doc = {
            "id": user_id,
            "username": user_data.username,
            "email": user_data.email,
            "first_name": user_data.first_name,
            "last_name": user_data.last_name,
            "role": user_data.role,
            "sub_role": user_data.sub_role,
            "department": user_data.department,
            "team": user_data.team,
            "pillar": user_data.pillar,
            "manager": user_data.manager,
            "approved_pillars": user_data.approved_pillars or [],
            "approved_departments": user_data.approved_departments or [],
            "created_at": datetime.now(timezone.utc).isoformat()
        }

        supabase.table("profiles").insert(profile_doc).execute()

        return User(
            id=str(user_id),
            username=user_data.username,
            email=user_data.email,
            first_name=user_data.first_name,
            last_name=user_data.last_name,
            role=user_data.role,
            sub_role=user_data.sub_role,
            department=user_data.department,
            team=user_data.team,
            pillar=user_data.pillar,
            manager=user_data.manager,
            approved_pillars=user_data.approved_pillars or [],
            approved_departments=user_data.approved_departments or [],
            created_at=profile_doc["created_at"]
        )
    except Exception as e:
        logging.error(f"Registration error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    profile = supabase.table("profiles").select("*").eq("username", credentials.username).maybeSingle().execute()
    if not profile.data:
        raise HTTPException(status_code=401, detail="Incorrect username or password")

    user_profile = profile.data

    try:
        auth_result = supabase.auth.sign_in_with_password({
            "email": user_profile["email"],
            "password": credentials.password
        })

        if not auth_result.user:
            raise HTTPException(status_code=401, detail="Incorrect username or password")

        token = create_access_token(data={"sub": str(user_profile["id"])})

        return TokenResponse(
            access_token=token,
            token_type="bearer",
            user=User(
                id=str(user_profile["id"]),
                username=user_profile["username"],
                email=user_profile["email"],
                first_name=user_profile.get("first_name"),
                last_name=user_profile.get("last_name"),
                role=user_profile["role"],
                sub_role=user_profile.get("sub_role"),
                department=user_profile.get("department"),
                team=user_profile.get("team"),
                pillar=user_profile.get("pillar"),
                manager=user_profile.get("manager"),
                approved_pillars=user_profile.get("approved_pillars") or [],
                approved_departments=user_profile.get("approved_departments") or [],
                created_at=user_profile["created_at"]
            )
        )
    except Exception as e:
        logging.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=401, detail="Incorrect username or password")


@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    return User(
        id=str(current_user["id"]),
        username=current_user["username"],
        email=current_user["email"],
        first_name=current_user.get("first_name"),
        last_name=current_user.get("last_name"),
        role=current_user["role"],
        sub_role=current_user.get("sub_role"),
        department=current_user.get("department"),
        team=current_user.get("team"),
        pillar=current_user.get("pillar"),
        manager=current_user.get("manager"),
        approved_pillars=current_user.get("approved_pillars") or [],
        approved_departments=current_user.get("approved_departments") or [],
        created_at=current_user["created_at"]
    )


@api_router.post("/auth/set-sub-role")
async def set_sub_role(selection: SubRoleSelection, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "approver":
        raise HTTPException(status_code=403, detail="Only approvers can set sub-role")

    if selection.sub_role not in ["approver", "ci_excellence"]:
        raise HTTPException(status_code=400, detail="Invalid sub-role")

    supabase.table("profiles").update({"sub_role": selection.sub_role}).eq("id", current_user["id"]).execute()

    return {"message": "Sub-role set successfully", "sub_role": selection.sub_role}


@api_router.post("/auth/change-password")
async def change_password(password_data: UserPasswordChange, current_user: dict = Depends(get_current_user)):
    try:
        supabase.auth.sign_in_with_password({
            "email": current_user["email"],
            "password": password_data.current_password
        })
    except Exception:
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    try:
        supabase.auth.admin.update_user_by_id(
            current_user["id"],
            {"password": password_data.new_password}
        )
        return {"message": "Password changed successfully"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    profile = supabase.table("profiles").select("*").eq("email", request.email).maybeSingle().execute()

    if not profile.data:
        return {"message": "If the email exists, a password reset link has been sent"}

    user = profile.data
    reset_token = create_reset_token(request.email)

    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"

    if RESEND_API_KEY:
        html = f"""
        <html>
            <body>
                <h2>Password Reset Request</h2>
                <p>Hello {user['username']},</p>
                <p>You requested to reset your password for Philtech Eye-dea.</p>
                <p>Click the link below to reset your password (valid for 1 hour):</p>
                <p><a href="{reset_link}">Reset Password</a></p>
                <p>If you didn't request this, please ignore this email.</p>
                <br>
                <p>Best regards,<br>Philtech Eye-dea Team</p>
            </body>
        </html>
        """
        asyncio.create_task(send_email_async(request.email, "Password Reset Request", html))
        return {
            "message": "Password reset link has been sent to your email",
            "note": "Using Resend test mode - emails only delivered to verified addresses",
            "reset_link": reset_link
        }
    else:
        return {
            "message": "Password reset link generated (email service not configured)",
            "reset_link": reset_link
        }


@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    email = verify_reset_token(request.token)

    profile = supabase.table("profiles").select("id").eq("email", email).maybeSingle().execute()
    if not profile.data:
        raise HTTPException(status_code=404, detail="User not found")

    try:
        supabase.auth.admin.update_user_by_id(
            profile.data["id"],
            {"password": request.new_password}
        )
        return {"message": "Password reset successfully. You can now login with your new password."}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def add_is_evaluated(idea_doc: dict) -> dict:
    idea_doc["is_evaluated"] = idea_doc.get("evaluated_by") is not None
    return idea_doc


def format_idea(idea: dict) -> Idea:
    idea = add_is_evaluated(idea)
    return Idea(
        id=str(idea["id"]),
        idea_number=idea["idea_number"],
        pillar=idea["pillar"],
        title=idea["title"],
        improvement_type=idea["improvement_type"],
        current_process=idea["current_process"],
        suggested_solution=idea["suggested_solution"],
        benefits=idea["benefits"],
        target_completion=idea.get("target_completion") or "",
        department=idea.get("department"),
        team=idea.get("team"),
        status=idea["status"],
        submitted_by=str(idea["submitted_by"]) if idea.get("submitted_by") else "",
        submitted_by_username=idea.get("submitted_by_username") or "",
        assigned_approver=str(idea["assigned_approver"]) if idea.get("assigned_approver") else None,
        assigned_approver_username=idea.get("assigned_approver_username"),
        created_at=idea["created_at"],
        updated_at=idea["updated_at"],
        is_quick_win=idea.get("is_quick_win"),
        complexity_level=idea.get("complexity_level"),
        savings_type=idea.get("savings_type"),
        cost_savings=float(idea["cost_savings"]) if idea.get("cost_savings") else None,
        time_saved_hours=float(idea["time_saved_hours"]) if idea.get("time_saved_hours") else None,
        time_saved_minutes=float(idea["time_saved_minutes"]) if idea.get("time_saved_minutes") else None,
        evaluation_notes=idea.get("evaluation_notes"),
        assigned_to_tech=idea.get("assigned_to_tech") or False,
        tech_person_name=idea.get("tech_person_name"),
        is_best_idea=idea.get("is_best_idea") or False,
        evaluated_by=str(idea["evaluated_by"]) if idea.get("evaluated_by") else None,
        evaluated_by_username=idea.get("evaluated_by_username"),
        evaluated_at=idea.get("evaluated_at"),
        is_evaluated=idea.get("is_evaluated") or False
    )


@api_router.get("/ideas", response_model=List[Idea])
async def get_ideas(
    status: Optional[str] = None,
    pillar: Optional[str] = None,
    department: Optional[str] = None,
    team: Optional[str] = None,
    submitted_by: Optional[str] = None,
    assigned_approver: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = supabase.table("ideas").select("*")
    if status:
        query = query.eq("status", status)
    if pillar:
        query = query.eq("pillar", pillar)
    if department:
        query = query.eq("department", department)
    if team:
        query = query.eq("team", team)
    if submitted_by:
        query = query.eq("submitted_by", submitted_by)
    if assigned_approver:
        query = query.eq("assigned_approver", assigned_approver)

    result = query.order("created_at", desc=True).execute()
    return [format_idea(idea) for idea in result.data]


@api_router.post("/ideas", response_model=Idea)
async def create_idea(idea_data: IdeaCreate, current_user: dict = Depends(get_current_user)):
    idea_number = await generate_idea_number()

    approver_query = supabase.table("profiles").select("*").eq("role", "approver")
    if idea_data.pillar:
        approver_query = approver_query.contains("approved_pillars", [idea_data.pillar])
    approver_result = approver_query.limit(1).execute()
    approver = approver_result.data[0] if approver_result.data else None

    if not approver and idea_data.department:
        approver_result = supabase.table("profiles").select("*").eq("role", "approver").eq("department", idea_data.department).limit(1).execute()
        approver = approver_result.data[0] if approver_result.data else None

    if not approver:
        approver_result = supabase.table("profiles").select("*").eq("role", "approver").limit(1).execute()
        approver = approver_result.data[0] if approver_result.data else None

    idea_doc = {
        "idea_number": idea_number,
        "pillar": idea_data.pillar,
        "title": idea_data.title,
        "improvement_type": idea_data.improvement_type,
        "current_process": idea_data.current_process,
        "suggested_solution": idea_data.suggested_solution,
        "benefits": idea_data.benefits,
        "target_completion": idea_data.target_completion,
        "department": idea_data.department,
        "team": idea_data.team,
        "status": "pending",
        "submitted_by": current_user["id"],
        "submitted_by_username": current_user["username"],
        "assigned_approver": approver["id"] if approver else None,
        "assigned_approver_username": approver["username"] if approver else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    result = supabase.table("ideas").insert(idea_doc).execute()
    created_idea = result.data[0]

    if approver:
        html = f"""
        <html>
            <body>
                <h2>New Eye-dea Submitted for Approval</h2>
                <p><strong>Idea Number:</strong> {idea_number}</p>
                <p><strong>Title:</strong> {idea_data.title}</p>
                <p><strong>Submitted By:</strong> {current_user['username']}</p>
                <p><strong>Pillar:</strong> {idea_data.pillar}</p>
                <p><strong>Department:</strong> {idea_data.department}</p>
                <p>Please review and approve/decline this Eye-dea.</p>
            </body>
        </html>
        """
        asyncio.create_task(send_email_async(approver["email"], f"New Eye-dea: {idea_data.title}", html))

    return format_idea(created_idea)


@api_router.get("/ideas/{idea_id}", response_model=Idea)
async def get_idea(idea_id: str, current_user: dict = Depends(get_current_user)):
    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")
    return format_idea(result.data)


@api_router.put("/ideas/{idea_id}", response_model=Idea)
async def update_idea(idea_id: str, idea_data: IdeaCreate, current_user: dict = Depends(get_current_user)):
    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data
    if str(idea["submitted_by"]) != str(current_user["id"]) and current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to update this idea")

    update_doc = {
        "pillar": idea_data.pillar,
        "title": idea_data.title,
        "improvement_type": idea_data.improvement_type,
        "current_process": idea_data.current_process,
        "suggested_solution": idea_data.suggested_solution,
        "benefits": idea_data.benefits,
        "target_completion": idea_data.target_completion,
        "department": idea_data.department,
        "team": idea_data.team,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    supabase.table("ideas").update(update_doc).eq("id", idea_id).execute()
    updated_result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    return format_idea(updated_result.data)


@api_router.delete("/ideas/{idea_id}")
async def delete_idea(idea_id: str, current_user: dict = Depends(get_admin_user)):
    result = supabase.table("ideas").delete().eq("id", idea_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")
    supabase.table("comments").delete().eq("idea_id", idea_id).execute()
    return {"message": "Idea deleted successfully"}


@api_router.post("/ideas/{idea_id}/approve")
async def approve_idea(idea_id: str, action: IdeaAction, current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "approver" and current_user.get("sub_role") == "ci_excellence":
        raise HTTPException(status_code=403, detail="C.I. Excellence Team cannot approve ideas. Only evaluate approved ideas.")
    if current_user["role"] not in ["approver", "admin"]:
        raise HTTPException(status_code=403, detail="Only approvers can approve ideas")

    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data

    supabase.table("ideas").update({
        "status": "approved",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    if action.comment:
        supabase.table("comments").insert({
            "idea_id": idea_id,
            "user_id": current_user["id"],
            "username": current_user["username"],
            "comment_text": action.comment,
            "created_at": datetime.now(timezone.utc).isoformat()
        }).execute()

    if idea.get("submitted_by"):
        submitter_result = supabase.table("profiles").select("*").eq("id", idea["submitted_by"]).maybeSingle().execute()
        if submitter_result.data:
            submitter = submitter_result.data
            html = f"""
            <html>
                <body>
                    <h2>Your Eye-dea Has Been Approved!</h2>
                    <p><strong>Idea Number:</strong> {idea['idea_number']}</p>
                    <p><strong>Title:</strong> {idea['title']}</p>
                    <p><strong>Approved By:</strong> {current_user['username']}</p>
                    {f'<p><strong>Comment:</strong> {action.comment}</p>' if action.comment else ''}
                </body>
            </html>
            """
            asyncio.create_task(send_email_async(submitter["email"], f"Eye-dea Approved: {idea['title']}", html))

    return {"message": "Idea approved successfully"}


@api_router.post("/ideas/{idea_id}/decline")
async def decline_idea(idea_id: str, action: IdeaAction, current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "approver" and current_user.get("sub_role") == "ci_excellence":
        raise HTTPException(status_code=403, detail="C.I. Excellence Team cannot decline ideas")
    if current_user["role"] not in ["approver", "admin"]:
        raise HTTPException(status_code=403, detail="Only approvers can decline ideas")

    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data

    supabase.table("ideas").update({
        "status": "declined",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    if action.comment:
        supabase.table("comments").insert({
            "idea_id": idea_id,
            "user_id": current_user["id"],
            "username": current_user["username"],
            "comment_text": action.comment,
            "created_at": datetime.now(timezone.utc).isoformat()
        }).execute()

    if idea.get("submitted_by"):
        submitter_result = supabase.table("profiles").select("*").eq("id", idea["submitted_by"]).maybeSingle().execute()
        if submitter_result.data:
            submitter = submitter_result.data
            html = f"""
            <html>
                <body>
                    <h2>Your Eye-dea Has Been Declined</h2>
                    <p><strong>Idea Number:</strong> {idea['idea_number']}</p>
                    <p><strong>Title:</strong> {idea['title']}</p>
                    <p><strong>Declined By:</strong> {current_user['username']}</p>
                    {f'<p><strong>Comment:</strong> {action.comment}</p>' if action.comment else ''}
                </body>
            </html>
            """
            asyncio.create_task(send_email_async(submitter["email"], f"Eye-dea Declined: {idea['title']}", html))

    return {"message": "Idea declined successfully"}


@api_router.post("/ideas/{idea_id}/request-revision")
async def request_revision(idea_id: str, action: IdeaAction, current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "approver" and current_user.get("sub_role") == "ci_excellence":
        raise HTTPException(status_code=403, detail="C.I. Excellence Team cannot request revisions")
    if current_user["role"] not in ["approver", "admin"]:
        raise HTTPException(status_code=403, detail="Only approvers can request revisions")

    if not action.comment:
        raise HTTPException(status_code=400, detail="Comment is required for revision requests")

    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data

    supabase.table("ideas").update({
        "status": "revision_requested",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    supabase.table("comments").insert({
        "idea_id": idea_id,
        "user_id": current_user["id"],
        "username": current_user["username"],
        "comment_text": action.comment,
        "created_at": datetime.now(timezone.utc).isoformat()
    }).execute()

    if idea.get("submitted_by"):
        submitter_result = supabase.table("profiles").select("*").eq("id", idea["submitted_by"]).maybeSingle().execute()
        if submitter_result.data:
            submitter = submitter_result.data
            html = f"""
            <html>
                <body>
                    <h2>Revision Requested for Your Eye-dea</h2>
                    <p><strong>Idea Number:</strong> {idea['idea_number']}</p>
                    <p><strong>Title:</strong> {idea['title']}</p>
                    <p><strong>Requested By:</strong> {current_user['username']}</p>
                    <p><strong>Comment:</strong> {action.comment}</p>
                    <p>Please revise and resubmit your Eye-dea.</p>
                </body>
            </html>
            """
            asyncio.create_task(send_email_async(submitter["email"], f"Revision Requested: {idea['title']}", html))

    return {"message": "Revision requested successfully"}


@api_router.post("/ideas/{idea_id}/resubmit")
async def resubmit_idea(idea_id: str, current_user: dict = Depends(get_current_user)):
    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data
    if str(idea["submitted_by"]) != str(current_user["id"]):
        raise HTTPException(status_code=403, detail="Not authorized to resubmit this idea")

    supabase.table("ideas").update({
        "status": "pending",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    if idea.get("assigned_approver"):
        approver_result = supabase.table("profiles").select("*").eq("id", idea["assigned_approver"]).maybeSingle().execute()
        if approver_result.data:
            approver = approver_result.data
            html = f"""
            <html>
                <body>
                    <h2>Eye-dea Resubmitted for Review</h2>
                    <p><strong>Idea Number:</strong> {idea['idea_number']}</p>
                    <p><strong>Title:</strong> {idea['title']}</p>
                    <p><strong>Submitted By:</strong> {current_user['username']}</p>
                    <p>This Eye-dea has been revised and resubmitted for your review.</p>
                </body>
            </html>
            """
            asyncio.create_task(send_email_async(approver["email"], f"Eye-dea Resubmitted: {idea['title']}", html))

    return {"message": "Idea resubmitted successfully"}


@api_router.get("/ideas/{idea_id}/comments", response_model=List[Comment])
async def get_comments(idea_id: str, current_user: dict = Depends(get_current_user)):
    result = supabase.table("comments").select("*").eq("idea_id", idea_id).order("created_at").execute()
    return [Comment(
        id=str(c["id"]),
        idea_id=str(c["idea_id"]),
        user_id=str(c["user_id"]),
        username=c["username"],
        comment_text=c["comment_text"],
        created_at=c["created_at"]
    ) for c in result.data]


@api_router.post("/ideas/{idea_id}/comments", response_model=Comment)
async def add_comment(idea_id: str, comment_data: CommentBase, current_user: dict = Depends(get_current_user)):
    idea_result = supabase.table("ideas").select("id").eq("id", idea_id).maybeSingle().execute()
    if not idea_result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    comment_doc = {
        "idea_id": idea_id,
        "user_id": current_user["id"],
        "username": current_user["username"],
        "comment_text": comment_data.comment_text,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    result = supabase.table("comments").insert(comment_doc).execute()
    created = result.data[0]
    return Comment(
        id=str(created["id"]),
        idea_id=str(created["idea_id"]),
        user_id=str(created["user_id"]),
        username=created["username"],
        comment_text=created["comment_text"],
        created_at=created["created_at"]
    )


@api_router.post("/ideas/{idea_id}/ci-evaluate")
async def ci_evaluate_idea(idea_id: str, evaluation: CIEvaluation, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "approver" or current_user.get("sub_role") != "ci_excellence":
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Only C.I. Excellence Team can evaluate ideas")

    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data

    new_status = idea.get("status", "approved")
    if evaluation.is_quick_win:
        new_status = "implemented"
    elif evaluation.assigned_to_tech and evaluation.tech_person_name:
        new_status = "assigned_to_te"

    update_doc = {
        "is_quick_win": evaluation.is_quick_win,
        "evaluated_by": current_user["id"],
        "evaluated_by_username": current_user["username"],
        "evaluated_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "status": new_status
    }

    if not evaluation.is_quick_win:
        update_doc["complexity_level"] = evaluation.complexity_level
        update_doc["savings_type"] = evaluation.savings_type
        update_doc["cost_savings"] = evaluation.cost_savings
        update_doc["time_saved_hours"] = evaluation.time_saved_hours
        update_doc["time_saved_minutes"] = evaluation.time_saved_minutes
        update_doc["evaluation_notes"] = evaluation.evaluation_notes
        update_doc["assigned_to_tech"] = evaluation.assigned_to_tech
        update_doc["tech_person_name"] = evaluation.tech_person_name

    supabase.table("ideas").update(update_doc).eq("id", idea_id).execute()

    if idea.get("submitted_by") and RESEND_API_KEY:
        submitter_result = supabase.table("profiles").select("*").eq("id", idea["submitted_by"]).maybeSingle().execute()
        if submitter_result.data:
            submitter = submitter_result.data
            html = f"""
            <html>
                <body>
                    <h2>Your Eye-dea Has Been Evaluated</h2>
                    <p><strong>Idea Number:</strong> {idea['idea_number']}</p>
                    <p><strong>Title:</strong> {idea['title']}</p>
                    <p><strong>Evaluated By:</strong> {current_user['username']} (C.I. Excellence Team)</p>
                    <p><strong>Quick Win:</strong> {'Yes' if evaluation.is_quick_win else 'No'}</p>
                    {f'<p><strong>Complexity Level:</strong> {evaluation.complexity_level}</p>' if not evaluation.is_quick_win else ''}
                </body>
            </html>
            """
            asyncio.create_task(send_email_async(submitter["email"], f"Eye-dea Evaluated: {idea['title']}", html))

    return {"message": "Idea evaluated successfully"}


@api_router.post("/ideas/{idea_id}/set-best-idea")
async def set_best_idea(idea_id: str, selection: BestIdeaSelection, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "approver" or current_user.get("sub_role") != "ci_excellence":
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Only C.I. Excellence Team can select best ideas")

    if selection.is_best_idea:
        supabase.table("ideas").update({"is_best_idea": False}).neq("id", idea_id).execute()

    supabase.table("ideas").update({
        "is_best_idea": selection.is_best_idea,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    return {"message": "Best idea status updated"}


@api_router.post("/ideas/{idea_id}/mark-best-idea")
async def mark_best_idea(idea_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "approver" or current_user.get("sub_role") != "ci_excellence":
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Only C.I. Excellence Team can select best ideas")

    result = supabase.table("ideas").select("id").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    supabase.table("ideas").update({"is_best_idea": False}).neq("id", idea_id).execute()

    supabase.table("ideas").update({
        "is_best_idea": True,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    return {"message": "Idea marked as best Eye-dea"}


class CIStatusUpdate(BaseModel):
    new_status: str


@api_router.post("/ideas/{idea_id}/ci-update-status")
async def ci_update_status(idea_id: str, status_update: CIStatusUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "approver" or current_user.get("sub_role") != "ci_excellence":
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Only C.I. Excellence Team can update idea status")

    result = supabase.table("ideas").select("*").eq("id", idea_id).maybeSingle().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Idea not found")

    idea = result.data
    if idea.get("status") != "assigned_to_te":
        raise HTTPException(status_code=400, detail="Can only change status of ideas assigned to T&E")

    valid_statuses = ["implemented", "revision_requested", "declined"]
    if status_update.new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

    supabase.table("ideas").update({
        "status": status_update.new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", idea_id).execute()

    return {"message": f"Idea status updated to {status_update.new_status}"}


@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    total_result = supabase.table("ideas").select("id", count="exact").execute()
    total = total_result.count or 0

    pending_result = supabase.table("ideas").select("id", count="exact").eq("status", "pending").execute()
    pending = pending_result.count or 0

    approved_result = supabase.table("ideas").select("id", count="exact").eq("status", "approved").execute()
    approved = approved_result.count or 0

    declined_result = supabase.table("ideas").select("id", count="exact").eq("status", "declined").execute()
    declined = declined_result.count or 0

    revision_result = supabase.table("ideas").select("id", count="exact").eq("status", "revision_requested").execute()
    revision = revision_result.count or 0

    my_ideas_result = supabase.table("ideas").select("id", count="exact").eq("submitted_by", current_user["id"]).execute()
    my_ideas = my_ideas_result.count or 0

    return {
        "total_ideas": total,
        "pending_ideas": pending,
        "approved_ideas": approved,
        "declined_ideas": declined,
        "revision_requested_ideas": revision,
        "my_ideas": my_ideas
    }


@api_router.get("/dashboard/analytics")
async def get_dashboard_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    base_query = supabase.table("ideas").select("*")
    if start_date:
        base_query = base_query.gte("created_at", start_date)
    if end_date:
        base_query = base_query.lte("created_at", end_date)

    all_ideas = base_query.execute()
    ideas_data = all_ideas.data

    total_ideas = len(ideas_data)
    declined_count = len([i for i in ideas_data if i.get("status") == "declined"])
    approved_count = len([i for i in ideas_data if i.get("status") == "approved"])
    implemented_count = len([i for i in ideas_data if i.get("status") == "implemented"])
    assigned_to_te_count = len([i for i in ideas_data if i.get("status") == "assigned_to_te"])
    pending_count = len([i for i in ideas_data if i.get("status") == "pending"])
    revision_count = len([i for i in ideas_data if i.get("status") == "revision_requested"])

    quick_wins_count = len([i for i in ideas_data if i.get("is_quick_win") == True])

    low_complexity = len([i for i in ideas_data if i.get("complexity_level") == "Low"])
    medium_complexity = len([i for i in ideas_data if i.get("complexity_level") == "Medium"])
    high_complexity = len([i for i in ideas_data if i.get("complexity_level") == "High"])

    best_idea_result = supabase.table("ideas").select("*").eq("is_best_idea", True).maybeSingle().execute()
    best_idea = best_idea_result.data

    total_cost_savings = sum(
        float(i.get("cost_savings") or 0)
        for i in ideas_data
        if i.get("savings_type") == "cost_savings" and i.get("cost_savings")
    )

    total_hours = sum(
        float(i.get("time_saved_hours") or 0)
        for i in ideas_data
        if i.get("savings_type") == "time_saved"
    )
    total_minutes = sum(
        float(i.get("time_saved_minutes") or 0)
        for i in ideas_data
        if i.get("savings_type") == "time_saved"
    )

    total_hours += int(total_minutes // 60)
    total_minutes = int(total_minutes % 60)

    denominator = total_ideas - declined_count
    approval_rate = (approved_count / denominator * 100) if denominator > 0 else 0
    implementation_rate = (implemented_count / denominator * 100) if denominator > 0 else 0
    assigned_to_te_rate = (assigned_to_te_count / denominator * 100) if denominator > 0 else 0

    return {
        "quick_wins_count": quick_wins_count,
        "complexity_counts": {
            "low": low_complexity,
            "medium": medium_complexity,
            "high": high_complexity
        },
        "best_idea": format_idea(best_idea) if best_idea else None,
        "total_cost_savings": total_cost_savings,
        "total_time_saved": {
            "hours": int(total_hours),
            "minutes": int(total_minutes)
        },
        "total_ideas": total_ideas,
        "approved_count": approved_count,
        "declined_count": declined_count,
        "implemented_count": implemented_count,
        "assigned_to_te_count": assigned_to_te_count,
        "pending_count": pending_count,
        "revision_count": revision_count,
        "approval_rate": round(approval_rate, 2),
        "implementation_rate": round(implementation_rate, 2),
        "assigned_to_te_rate": round(assigned_to_te_rate, 2),
        "charts_data": {
            "complexity_chart": [
                {"name": "Low Complexity", "value": low_complexity},
                {"name": "Medium Complexity", "value": medium_complexity},
                {"name": "High Complexity", "value": high_complexity}
            ],
            "quick_wins_chart": [
                {"name": "Quick Wins", "value": quick_wins_count},
                {"name": "Not Quick Wins", "value": low_complexity + medium_complexity + high_complexity}
            ],
            "status_chart": [
                {"name": "Approved", "value": approved_count},
                {"name": "Implemented", "value": implemented_count},
                {"name": "Assigned to T&E", "value": assigned_to_te_count},
                {"name": "Pending", "value": pending_count},
                {"name": "Revision Requested", "value": revision_count},
                {"name": "Declined", "value": declined_count}
            ]
        }
    }


@api_router.get("/dashboard/export-excel")
async def export_ideas_excel(current_user: dict = Depends(get_current_user)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    result = supabase.table("ideas").select("*").execute()
    ideas = result.data

    wb = Workbook()
    ws = wb.active
    ws.title = "Eye-deas"

    headers = [
        "Idea Number", "Title", "Status", "Pillar", "Department", "Team",
        "Improvement Type", "Submitted By", "Assigned Approver",
        "Quick Win", "Complexity", "Savings Type", "Cost Savings",
        "Time Saved (Hours)", "Time Saved (Minutes)", "Evaluated By",
        "Tech Person", "Best Idea", "Target Completion", "Created At"
    ]

    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, idea in enumerate(ideas, 2):
        ws.cell(row=row_num, column=1, value=idea.get("idea_number"))
        ws.cell(row=row_num, column=2, value=idea.get("title"))
        ws.cell(row=row_num, column=3, value=idea.get("status"))
        ws.cell(row=row_num, column=4, value=idea.get("pillar"))
        ws.cell(row=row_num, column=5, value=idea.get("department"))
        ws.cell(row=row_num, column=6, value=idea.get("team"))
        ws.cell(row=row_num, column=7, value=idea.get("improvement_type"))
        ws.cell(row=row_num, column=8, value=idea.get("submitted_by_username"))
        ws.cell(row=row_num, column=9, value=idea.get("assigned_approver_username"))
        ws.cell(row=row_num, column=10, value="Yes" if idea.get("is_quick_win") else "No" if idea.get("is_quick_win") is not None else "")
        ws.cell(row=row_num, column=11, value=idea.get("complexity_level") or "")
        ws.cell(row=row_num, column=12, value=idea.get("savings_type") or "")
        ws.cell(row=row_num, column=13, value=idea.get("cost_savings") or "")
        ws.cell(row=row_num, column=14, value=idea.get("time_saved_hours") or "")
        ws.cell(row=row_num, column=15, value=idea.get("time_saved_minutes") or "")
        ws.cell(row=row_num, column=16, value=idea.get("evaluated_by_username") or "")
        ws.cell(row=row_num, column=17, value=idea.get("tech_person_name") or "")
        ws.cell(row=row_num, column=18, value="Yes" if idea.get("is_best_idea") else "No")
        ws.cell(row=row_num, column=19, value=idea.get("target_completion"))
        ws.cell(row=row_num, column=20, value=idea.get("created_at"))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=philtech_eyedeas.xlsx"}
    )


@api_router.get("/admin/users", response_model=List[User])
async def get_users(current_user: dict = Depends(get_admin_user)):
    demo_usernames = ["admin", "approver1", "user1"]
    result = supabase.table("profiles").select("*").execute()
    users = [u for u in result.data if u["username"] not in demo_usernames]
    return [User(
        id=str(u["id"]),
        username=u["username"],
        email=u["email"],
        first_name=u.get("first_name"),
        last_name=u.get("last_name"),
        role=u["role"],
        sub_role=u.get("sub_role"),
        department=u.get("department"),
        team=u.get("team"),
        pillar=u.get("pillar"),
        manager=u.get("manager"),
        approved_pillars=u.get("approved_pillars") or [],
        approved_departments=u.get("approved_departments") or [],
        created_at=u["created_at"]
    ) for u in users]


@api_router.put("/admin/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_data: UserBase, current_user: dict = Depends(get_admin_user)):
    update_doc = {
        "username": user_data.username,
        "email": user_data.email,
        "role": user_data.role,
        "department": user_data.department,
        "team": user_data.team,
        "pillar": user_data.pillar,
        "manager": user_data.manager,
        "approved_pillars": user_data.approved_pillars if user_data.role == "approver" else [],
        "approved_departments": user_data.approved_departments if user_data.role == "approver" else []
    }

    result = supabase.table("profiles").update(update_doc).eq("id", user_id).execute()

    if not result.data:
        raise HTTPException(status_code=404, detail="User not found")

    updated = result.data[0]
    return User(
        id=str(updated["id"]),
        username=updated["username"],
        email=updated["email"],
        first_name=updated.get("first_name"),
        last_name=updated.get("last_name"),
        role=updated["role"],
        sub_role=updated.get("sub_role"),
        department=updated.get("department"),
        team=updated.get("team"),
        pillar=updated.get("pillar"),
        manager=updated.get("manager"),
        approved_pillars=updated.get("approved_pillars") or [],
        approved_departments=updated.get("approved_departments") or [],
        created_at=updated["created_at"]
    )


@api_router.post("/admin/users/bulk-upload")
async def bulk_upload_users(file: bytes = File(...), current_user: dict = Depends(get_admin_user)):
    import csv
    import io

    try:
        csv_content = file.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))

        created_users = []
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                if not row.get('username') or not row.get('email') or not row.get('password'):
                    errors.append(f"Row {row_num}: Missing required fields (username, email, password)")
                    continue

                existing = supabase.table("profiles").select("id").eq("username", row['username']).maybeSingle().execute()
                if existing.data:
                    errors.append(f"Row {row_num}: Username '{row['username']}' already exists")
                    continue

                auth_result = supabase.auth.sign_up({
                    "email": row['email'],
                    "password": row['password'],
                    "options": {
                        "data": {
                            "username": row['username'],
                            "role": row.get('role', 'user')
                        }
                    }
                })

                if not auth_result.user:
                    errors.append(f"Row {row_num}: Failed to create auth user")
                    continue

                approved_pillars = row.get('approved_pillars', '').split(';') if row.get('approved_pillars') else []
                approved_departments = row.get('approved_departments', '').split(';') if row.get('approved_departments') else []

                profile_doc = {
                    "id": auth_result.user.id,
                    "username": row['username'],
                    "email": row['email'],
                    "role": row.get('role', 'user'),
                    "department": row.get('department', ''),
                    "team": row.get('team', ''),
                    "pillar": row.get('pillar', ''),
                    "manager": row.get('manager', ''),
                    "approved_pillars": approved_pillars,
                    "approved_departments": approved_departments,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }

                supabase.table("profiles").insert(profile_doc).execute()
                created_users.append(row['username'])
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        return {
            "message": f"Bulk upload completed. Created {len(created_users)} users.",
            "created_users": created_users,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_admin_user)):
    user_result = supabase.table("profiles").select("username").eq("id", user_id).maybeSingle().execute()
    if user_result.data and user_result.data.get("username") in ["admin", "approver1", "user1"]:
        raise HTTPException(status_code=403, detail="Cannot delete demo accounts")

    result = supabase.table("profiles").delete().eq("id", user_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}


@api_router.get("/admin/departments", response_model=List[Department])
async def get_departments(pillar: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = supabase.table("departments").select("*")
    if pillar:
        query = query.eq("pillar", pillar)
    result = query.execute()
    return [Department(id=str(d["id"]), name=d["name"], pillar=d["pillar"]) for d in result.data]


@api_router.post("/admin/departments", response_model=Department)
async def create_department(dept_data: DepartmentBase, current_user: dict = Depends(get_admin_user)):
    dept_doc = {"name": dept_data.name, "pillar": dept_data.pillar}
    result = supabase.table("departments").insert(dept_doc).execute()
    created = result.data[0]
    return Department(id=str(created["id"]), name=created["name"], pillar=created["pillar"])


@api_router.delete("/admin/departments/{dept_id}")
async def delete_department(dept_id: str, current_user: dict = Depends(get_admin_user)):
    result = supabase.table("departments").delete().eq("id", dept_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"message": "Department deleted successfully"}


@api_router.get("/admin/pillars", response_model=List[Pillar])
async def get_pillars(current_user: dict = Depends(get_current_user)):
    result = supabase.table("pillars").select("*").execute()
    return [Pillar(id=str(p["id"]), name=p["name"]) for p in result.data]


@api_router.post("/admin/pillars", response_model=Pillar)
async def create_pillar(pillar_data: PillarBase, current_user: dict = Depends(get_admin_user)):
    pillar_doc = {"name": pillar_data.name}
    result = supabase.table("pillars").insert(pillar_doc).execute()
    created = result.data[0]
    return Pillar(id=str(created["id"]), name=created["name"])


@api_router.delete("/admin/pillars/{pillar_id}")
async def delete_pillar(pillar_id: str, current_user: dict = Depends(get_admin_user)):
    result = supabase.table("pillars").delete().eq("id", pillar_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Pillar not found")
    return {"message": "Pillar deleted successfully"}


@api_router.get("/admin/teams", response_model=List[Team])
async def get_teams(pillar: Optional[str] = None, department: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = supabase.table("teams").select("*")
    if pillar:
        query = query.eq("pillar", pillar)
    if department:
        query = query.eq("department", department)
    result = query.execute()
    return [Team(id=str(t["id"]), name=t["name"], pillar=t["pillar"], department=t["department"]) for t in result.data]


@api_router.post("/admin/teams", response_model=Team)
async def create_team(team_data: TeamBase, current_user: dict = Depends(get_admin_user)):
    team_doc = {"name": team_data.name, "pillar": team_data.pillar, "department": team_data.department}
    result = supabase.table("teams").insert(team_doc).execute()
    created = result.data[0]
    return Team(id=str(created["id"]), name=created["name"], pillar=created["pillar"], department=created["department"])


@api_router.delete("/admin/teams/{team_id}")
async def delete_team(team_id: str, current_user: dict = Depends(get_admin_user)):
    result = supabase.table("teams").delete().eq("id", team_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Team not found")
    return {"message": "Team deleted successfully"}


@api_router.get("/admin/tech-persons", response_model=List[TechPerson])
async def get_tech_persons(current_user: dict = Depends(get_current_user)):
    result = supabase.table("tech_persons").select("*").execute()
    return [TechPerson(
        id=str(p["id"]),
        name=p["name"],
        email=p.get("email"),
        specialization=p.get("specialization")
    ) for p in result.data]


@api_router.post("/admin/tech-persons", response_model=TechPerson)
async def create_tech_person(person_data: TechPersonBase, current_user: dict = Depends(get_admin_user)):
    person_doc = {
        "name": person_data.name,
        "email": person_data.email,
        "specialization": person_data.specialization
    }
    result = supabase.table("tech_persons").insert(person_doc).execute()
    created = result.data[0]
    return TechPerson(
        id=str(created["id"]),
        name=created["name"],
        email=created.get("email"),
        specialization=created.get("specialization")
    )


@api_router.delete("/admin/tech-persons/{person_id}")
async def delete_tech_person(person_id: str, current_user: dict = Depends(get_admin_user)):
    result = supabase.table("tech_persons").delete().eq("id", person_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Tech person not found")
    return {"message": "Tech person deleted successfully"}


@api_router.post("/admin/seed-data")
async def seed_data(current_user: dict = Depends(get_admin_user)):
    existing = supabase.table("pillars").select("id", count="exact").execute()
    if existing.count and existing.count > 0:
        return {"message": "Data already seeded"}

    pillars = ["GBS", "Tech", "Finance", "HR"]
    for pillar_name in pillars:
        supabase.table("pillars").insert({"name": pillar_name}).execute()

    departments = [
        {"name": "Operations", "pillar": "GBS"},
        {"name": "Technology", "pillar": "Tech"},
        {"name": "Finance", "pillar": "Finance"},
        {"name": "Human Resources", "pillar": "HR"}
    ]
    for dept in departments:
        supabase.table("departments").insert(dept).execute()

    teams = [
        {"name": "Allowance Billing", "pillar": "GBS", "department": "Operations"},
        {"name": "Pre-audit and AB", "pillar": "GBS", "department": "Operations"}
    ]
    for team in teams:
        supabase.table("teams").insert(team).execute()

    return {"message": "Sample data seeded successfully"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
